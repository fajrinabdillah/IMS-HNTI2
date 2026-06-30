#!/usr/bin/env python3
"""Convert HNTI sales weekly Excel files to IMS Field Report CSV import format."""

from __future__ import annotations

import csv
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

DESKTOP = Path("/Users/fajrin2/Desktop")
OUT_DIR = DESKTOP

HEADER = [
    "Report ID", "Sales ID", "Sales Name", "Date", "Week", "Field Days", "Nights", "Focus Area",
    "Pipeline RS Count", "Pipeline Value (Rp juta)", "Closest RS", "Win", "Blocker", "Next Priority", "Fatigue",
    "Hospital", "City", "Visit Type", "Product", "Pipeline Temp", "Contact", "Visit Note",
    "Estimasi Nilai Proyek (Rp juta)",
]

SALES = {
    "hatim": ("hatim", "Ahmad Hatim Ashshidiq", "Jateng Utara & Pantura"),
    "dwi": ("dwi", "Dwi Wahyudianto", "Jabodetabek + Banten + Jabar"),
    "tri": ("tri", "Tri Sutjahjono", "Jatim Selatan & Timur"),
    "bagus": ("bagus", "Bagus Iswahyudi", "Jatim Utara & Barat"),
    "icha": ("icha", "Ika Apriani", "Jabodetabek & Banten"),
    "astrika": ("astrika", "Astrika Sari", "Jateng Selatan & DIY"),
}

PRODUCT_KEYWORDS = [
    (r"ct\s*scan|ctscan|\bct\b", "CT Scan", 9000),
    (r"mri|mr\s*1\.?5", "MRI", 18000),
    (r"c-?arm|c arm", "C-Arm", 4000),
    (r"mammo", "Mammo", 2000),
    (r"eswl", "ESWL", 3000),
    (r"portable\s*x-?ray|mobile\s*x-?ray|mobile\s*dr", "X-Ray", 600),
    (r"x-?ray|stationery|stationary|ceiling", "X-Ray", 900),
    (r"fpd", "X-Ray", 700),
    (r"endoscop", "Other", 1500),
    (r"usg", "Other", 800),
]

PIPELINE_FROM_TEXT = [
    (r"closing|deal|po|menang|closed", "win", 1.0),
    (r"sph|proposal|penawaran|beauty contest|presentasi", "proposal", 0.85),
    (r"hot|urgent|segera|butuh|kebutuhan", "hot", 0.7),
    (r"follow\s*up|wacana|rencana|tertarik|investigasi", "warm", 0.5),
    (r"tidak ada kebutuhan|belum|kosong|off day|libur|cuti", "cold", 0.2),
]

VISIT_FROM_TEXT = [
    (r"closing|deal|menang", "closed"),
    (r"nego|negosiasi|harga", "nego"),
    (r"demo|trial", "demo"),
    (r"follow\s*up|followup", "followup"),
]


def norm_date(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})", s, re.I)
    if m:
        months = {
            "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
            "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "agu": 8, "sep": 9, "okt": 10, "nov": 11, "des": 12,
        }
        mo = months.get(m.group(2).lower()[:3], months.get(m.group(2).lower()))
        if mo:
            return f"{int(m.group(3)):04d}-{mo:02d}-{int(m.group(1)):02d}"
    return ""


def parse_money_juta(text: str) -> int:
    if not text:
        return 0
    t = text.lower().replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)\s*(jt|juta|m\b|milyar|b\b)", t)
    if m:
        val = float(m.group(1))
        if m.group(2) in ("milyar", "b"):
            return int(val * 1000)
        return int(val)
    m = re.search(r"under\s*(\d+(?:\.\d+)?)\s*m", t)
    if m:
        return int(float(m.group(1)) * 0.8)
    return 0


def detect_product(text: str) -> tuple[str, int]:
    t = (text or "").lower()
    for pat, prod, base in PRODUCT_KEYWORDS:
        if re.search(pat, t):
            return prod, base
    return "", 0


def detect_pipeline(text: str) -> tuple[str, float]:
    t = (text or "").lower()
    for pat, temp, mult in PIPELINE_FROM_TEXT:
        if re.search(pat, t):
            return temp, mult
    return "warm", 0.5


def detect_visit_type(text: str, pipeline: str) -> str:
    t = (text or "").lower()
    for pat, vt in VISIT_FROM_TEXT:
        if re.search(pat, t):
            return vt
    if pipeline == "win":
        return "closed"
    if pipeline in ("proposal", "hot"):
        return "followup"
    return "first"


def estimate_value(text: str, pipeline: str, product: str, base: int) -> int:
    explicit = parse_money_juta(text)
    if explicit:
        return explicit
    if not base:
        _, base = detect_product(text)
    if not base:
        defaults = {"CT Scan": 9000, "MRI": 18000, "C-Arm": 4000, "Mammo": 2000, "ESWL": 3000, "X-Ray": 800}
        base = defaults.get(product, 0)
    if not base:
        return 0
    _, mult = detect_pipeline(text)
    if pipeline in ("win", "proposal", "hot", "warm", "cold"):
        pmap = {"win": 1.0, "proposal": 0.85, "hot": 0.7, "warm": 0.5, "cold": 0.25}
        mult = pmap.get(pipeline, mult)
    return max(int(base * mult), 0)


def week_label(date_str: str) -> str:
    if not date_str:
        return "Minggu 1"
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return f"Minggu {min((d.day - 1) // 7 + 1, 4)}"


def make_row(report_meta, visit, report_id):
    sid, sname, _ = SALES[report_meta["sales_key"]]
    return [
        report_id,
        sid,
        sname,
        visit.get("date") or report_meta.get("end_date") or report_meta.get("start_date", ""),
        report_meta.get("week", week_label(visit.get("date", ""))),
        report_meta.get("days", 5),
        report_meta.get("nights", 3),
        report_meta.get("area", ""),
        "",  # pipeN filled later
        "",  # pipeVal filled later
        report_meta.get("closest", ""),
        report_meta.get("win", ""),
        report_meta.get("block", ""),
        report_meta.get("next", ""),
        report_meta.get("fatigue", 2),
        visit.get("name", ""),
        visit.get("city", ""),
        visit.get("visit", "first"),
        visit.get("product", ""),
        visit.get("pipeline", "warm"),
        visit.get("contact", ""),
        visit.get("note", ""),
        visit.get("estimatedValue", 0),
    ]


def finalize_rows(rows: list[list]) -> list[list]:
    """Fill pipeN/pipeVal per report id."""
    by_rid: dict[str, list[list]] = {}
    for row in rows:
        by_rid.setdefault(row[0], []).append(row)
    out = []
    for rid, grp in by_rid.items():
        hot_warm = sum(1 for r in grp if r[19] in ("hot", "warm", "proposal", "win"))
        total_val = sum(int(r[22] or 0) for r in grp)
        for r in grp:
            r[8] = hot_warm
            r[9] = total_val
            out.append(r)
    return out


def write_csv(path: Path, rows: list[list]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        w.writerows(rows)


def is_hospital_name(s: str) -> bool:
    s = s.strip()
    if not s or len(s) < 4:
        return False
    upper = s.upper()
    if upper in ("JATENG", "JAKARTA BARAT", "TANGERANG", "CIPUTAT", "TANGERANG SELATAN", "SEMARANG", "KANTOR", "WFH", "LIBUR PENGGANTI DINAS"):
        return False
    if re.match(r"^(JV|MEETING|OFF DAY|CUTI|LIBUR|EVENT|JAKARTA|PROBOLINGGO|GRESIK|KRIAN|JOMBANG|MOJOKERTO)\b", upper):
        return False
    return bool(re.search(r"\b(RS|RSIA|RSUD|RSAL|RSU|Klinik|Hospital|PT)\b", s, re.I)) or s.upper().startswith("RS")


def parse_hatim_table(path: Path, is_plan: bool) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    title = str(rows[0][0] or "")
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4}).*?(\d{1,2})/(\d{1,2})/(\d{2,4})", title)
    year = 2026
    if m:
        start = norm_date(f"{year}-{int(m.group(2)):02d}-{int(m.group(1)):02d}")
        end = norm_date(f"{year}-{int(m.group(5)):02d}-{int(m.group(4)):02d}")
    else:
        start, end = "2026-06-22", "2026-06-26"

    rid = f"rpt_hatim_{start.replace('-', '')}_{'plan' if is_plan else 'report'}"
    meta = {
        "sales_key": "hatim",
        "start_date": start,
        "end_date": end,
        "week": week_label(end),
        "days": 5,
        "nights": 3,
        "area": "Jateng Utara & Pantura",
        "win": "" if is_plan else "Follow up penawaran C-Arm & CT trade-in RS Banyumanik 2",
        "block": "" if is_plan else "Vendor harus terdaftar formularium Hermina pusat",
        "next": "Visit RS Charlie Demak, Sultan Agung, Pantiwilasa" if is_plan else "Penawaran X-Ray RS Banyumanik 1, follow up C-Arm",
    }

    out_rows = []
    current_date = start
    for row in rows[2:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if cells[0].isdigit():
            current_date = norm_date(cells[1]) or current_date
        outlet = cells[2] if len(cells) > 2 else ""
        cp = cells[3] if len(cells) > 3 else ""
        note = cells[4] if len(cells) > 4 else ""

        if outlet.upper().startswith("JV DENGAN"):
            # Split bundled narrative visits
            blocks = re.split(r"\n\s*\n|(?=Rs\s)", note, flags=re.I)
            for block in blocks:
                block = block.strip()
                if not block:
                    continue
                hm = re.match(r"(?:Rs\.?\s*)?(.+?)(?:\n|$)", block, re.I)
                if not hm:
                    continue
                name = hm.group(1).strip().split("\n")[0]
                if not is_hospital_name(name):
                    continue
                product, base = detect_product(block)
                pipeline, _ = detect_pipeline(block)
                val = estimate_value(block, pipeline, product, base)
                out_rows.append(make_row(meta, {
                    "date": current_date,
                    "name": name if name.upper().startswith("RS") else f"RS {name}",
                    "city": "",
                    "visit": detect_visit_type(block, pipeline),
                    "product": product,
                    "pipeline": pipeline,
                    "contact": cp,
                    "note": block[:500],
                    "estimatedValue": val,
                }, rid))
            continue

        if not is_hospital_name(outlet):
            continue
        product, base = detect_product(note or outlet)
        pipeline, _ = detect_pipeline(note)
        if is_plan and not note:
            pipeline = "warm"
        val = estimate_value(note, pipeline, product, base)
        out_rows.append(make_row(meta, {
            "date": current_date,
            "name": outlet,
            "city": "",
            "visit": "first" if is_plan else detect_visit_type(note, pipeline),
            "product": product,
            "pipeline": pipeline,
            "contact": cp,
            "note": note or ("Rencana kunjungan" if is_plan else ""),
            "estimatedValue": val if val else (3000 if is_plan else 0),
        }, rid))
    return finalize_rows(out_rows)


def parse_tri_report(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    rid = "rpt_tri_20260626_report"
    meta = {
        "sales_key": "tri",
        "start_date": "2026-06-22",
        "end_date": "2026-06-26",
        "week": "Minggu 4",
        "days": 5,
        "nights": 3,
        "area": "Jatim Selatan & Timur (Surabaya-Sidoarjo)",
        "win": "Mapping kebutuhan RSAL & RS Mitra Waru",
        "block": "Pimpinan RSAL dinas luar, belum bisa dihubungi",
        "next": "Follow up program anggaran RSAL",
    }
    out_rows = []
    current_date = "2026-06-22"
    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells or not any(cells):
            continue
        if re.match(r"^\d+$", cells[0]) and norm_date(cells[1]):
            current_date = norm_date(cells[1])
            continue
        text = " ".join(c for c in cells if c)
        if len(text) < 20:
            continue
        hm = re.search(r"(RS\s*[A-Za-z0-9 .\-]+?)(?:\s+visit|\s+:|\s+-->)", text, re.I)
        name = hm.group(1).strip() if hm else ""
        if not name:
            if "RSAL" in text.upper():
                name = "RSAL dr. Ramelan"
            elif "RS Mitra Waru" in text:
                name = "RS Mitra Waru"
            else:
                continue
        product, base = detect_product(text)
        pipeline, _ = detect_pipeline(text)
        val = estimate_value(text, pipeline, product, base)
        out_rows.append(make_row(meta, {
            "date": current_date,
            "name": name,
            "city": "",
            "visit": detect_visit_type(text, pipeline),
            "product": product,
            "pipeline": pipeline,
            "contact": "",
            "note": text[:500],
            "estimatedValue": val,
        }, rid))
    return finalize_rows(out_rows)


def parse_tri_plan(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = [[str(c).strip() if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    dates = rows[4][2:7] if len(rows) > 4 else []
    rid = "rpt_tri_20260703_plan"
    meta = {
        "sales_key": "tri",
        "start_date": "2026-06-29",
        "end_date": "2026-07-03",
        "week": "Minggu 1",
        "days": 3,
        "nights": 2,
        "area": "Jatim Selatan & Timur",
        "next": "Visit RS Panti Waluya, Suyudi Lamongan, RSUD Haryoto Lumajang",
    }
    out_rows = []
    for r in rows[5:]:
        if not any(r):
            continue
        for col_idx, d_raw in enumerate(dates, start=2):
            if col_idx >= len(r):
                break
            name = r[col_idx]
            if not is_hospital_name(name):
                continue
            d = norm_date(d_raw)
            product, base = detect_product(name)
            out_rows.append(make_row(meta, {
                "date": d,
                "name": name,
                "city": "",
                "visit": "first",
                "product": product,
                "pipeline": "warm",
                "contact": "",
                "note": "Rencana kunjungan mingguan",
                "estimatedValue": estimate_value(name, "warm", product, base) or 5000,
            }, rid))
    return finalize_rows(out_rows)


def parse_ika_plan(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    rid = "rpt_icha_20260703_plan"
    meta = {
        "sales_key": "icha",
        "start_date": "2026-06-29",
        "end_date": "2026-07-03",
        "week": "Minggu 1",
        "days": 4,
        "nights": 2,
        "area": "Jabodetabek & Banten",
        "next": "Kunjungan RS Jabodetabek W1 Juli",
    }
    out_rows = []
    current_date = "2026-06-29"
    current_city = ""
    for row in rows[8:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if cells[1].isdigit() and norm_date(cells[2]):
            current_date = norm_date(cells[2])
            city_candidate = cells[3]
            if city_candidate and not is_hospital_name(city_candidate):
                current_city = city_candidate
            continue
        name = cells[3] if len(cells) > 3 else ""
        if not is_hospital_name(name):
            continue
        product, base = detect_product(name)
        out_rows.append(make_row(meta, {
            "date": current_date,
            "name": name,
            "city": current_city,
            "visit": "first",
            "product": product,
            "pipeline": "warm",
            "contact": "",
            "note": "Rencana kunjungan",
            "estimatedValue": estimate_value(name, "warm", product, base) or 6000,
        }, rid))
    return finalize_rows(out_rows)


def parse_dwi_ika_report(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["WEEK 4 "] if "WEEK 4 " in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    rid = "rpt_dwi_20260623_jateng"
    meta = {
        "sales_key": "dwi",
        "start_date": "2026-06-23",
        "end_date": "2026-06-23",
        "week": "Minggu 4",
        "days": 1,
        "nights": 0,
        "area": "Jawa Tengah (joint visit)",
        "win": "Joint visit Jateng dengan Hatim",
        "next": "Follow up RS Banyumanik presentasi",
    }
    out_rows = []
    current_city = ""
    for row in rows[4:]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        if cells[1].isdigit():
            city_candidate = cells[3]
            if city_candidate and not is_hospital_name(city_candidate):
                current_city = city_candidate
            continue
        name = cells[3]
        if not is_hospital_name(name):
            continue
        note = cells[6] if len(cells) > 6 else ""
        contact = cells[4] if len(cells) > 4 else ""
        product, base = detect_product(note + " " + name)
        pipeline, _ = detect_pipeline(note)
        val = estimate_value(note, pipeline, product, base)
        out_rows.append(make_row(meta, {
            "date": "2026-06-23",
            "name": name if name.upper().startswith("RS") else f"RS {name}",
            "city": current_city,
            "visit": detect_visit_type(note, pipeline),
            "product": product,
            "pipeline": pipeline,
            "contact": contact,
            "note": note,
            "estimatedValue": val,
        }, rid))
    return finalize_rows(out_rows)


def parse_dwi_plan(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = [[str(c).strip() if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    rid = "rpt_dwi_20260703_plan"
    meta = {
        "sales_key": "dwi",
        "start_date": "2026-06-29",
        "end_date": "2026-07-03",
        "week": "Minggu 1",
        "days": 5,
        "nights": 2,
        "area": "Jabodetabek + Banten + Jabar",
        "next": "Meeting RS Bhakti Kartini, visit Bogor & Tangerang",
    }
    dates = [norm_date(c) for c in rows[12][0:5]] if len(rows) > 12 else []
    out_rows = []
    for r in rows[13:16]:
        for col_idx, d in enumerate(dates):
            if col_idx >= len(r):
                break
            name = r[col_idx]
            if not is_hospital_name(name) and not name.upper().startswith("MEETING"):
                continue
            if name.upper().startswith("MEETING"):
                name = name.replace("Meeting", "Meeting").strip()
            if name.upper() == "KANTOR":
                continue
            product, base = detect_product(name)
            out_rows.append(make_row(meta, {
                "date": d,
                "name": name if is_hospital_name(name) else name,
                "city": "",
                "visit": "first",
                "product": product,
                "pipeline": "warm",
                "contact": "",
                "note": "Rencana minggu 29 Jun - 3 Jul",
                "estimatedValue": estimate_value(name, "warm", product, base) or 8000,
            }, rid))
    return finalize_rows(out_rows)


def parse_bagus_sheet(ws, sheet_name: str) -> list[list]:
    rows = [[str(c).strip() if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    if sheet_name != "2906 sd 0307":
        return []

    rid = "rpt_bagus_20260703_plan"
    meta = {
        "sales_key": "bagus",
        "start_date": "2026-06-29",
        "end_date": "2026-07-03",
        "week": "Minggu 1",
        "days": 5,
        "nights": 3,
        "area": "Jatim Utara & Barat",
        "next": "Follow up pipeline Jatim 2 pasca IHEX",
    }
    out_rows = []

    # Plan block rows 148+
    dates = []
    if len(rows) > 148:
        for c in rows[148][1:6]:
            if isinstance(c, str) and c.isdigit():
                dt = datetime(1899, 12, 30) + timedelta(days=int(c))
                dates.append(dt.strftime("%Y-%m-%d"))
    if not dates:
        dates = ["2026-06-29", "2026-06-30", "2026-07-01", "2026-07-02", "2026-07-03"]

    for r in rows[149:154]:
        for col_idx in range(min(5, len(r))):
            name = r[col_idx]
            if not is_hospital_name(name):
                continue
            d = dates[col_idx] if col_idx < len(dates) else dates[0]
            product, base = detect_product(name)
            out_rows.append(make_row(meta, {
                "date": d,
                "name": name,
                "city": "",
                "visit": "first",
                "product": product,
                "pipeline": "warm",
                "contact": "",
                "note": "Rencana kunjungan Jatim 2",
                "estimatedValue": estimate_value(name, "warm", product, base) or 7000,
            }, rid))

    # Pipeline table 162+
    rid2 = "rpt_bagus_20260626_pipeline"
    meta2 = {
        **meta,
        "start_date": "2026-06-22",
        "end_date": "2026-06-26",
        "week": "Minggu 4",
        "win": "Pipeline Jatim 2 aktif pasca IHEX",
    }
    for r in rows[162:180]:
        if len(r) < 4 or not r[0].upper().startswith("RS"):
            continue
        name = r[0]
        need = r[1]
        stage = r[2]
        nxt = r[3]
        note = f"{need}. {stage}. {nxt}"
        product, base = detect_product(need)
        pipeline, _ = detect_pipeline(stage + " " + nxt)
        if "sph" in stage.lower() or "sph" in nxt.lower():
            pipeline = "proposal"
        val = estimate_value(note, pipeline, product, base)
        out_rows.append(make_row(meta2, {
            "date": "2026-06-26",
            "name": name,
            "city": "",
            "visit": detect_visit_type(note, pipeline),
            "product": product,
            "pipeline": pipeline,
            "contact": "",
            "note": note[:500],
            "estimatedValue": val,
        }, rid2))

    # Realisasi 22-26 Jun block
    rid3 = "rpt_bagus_20260626_report"
    meta3 = {**meta, "start_date": "2026-06-22", "end_date": "2026-06-26", "week": "Minggu 4"}
    result_rows = [i for i, r in enumerate(rows) if r and r[0] == "Result :" and any("RS" in c for c in r)]
    for ri in result_rows:
        if ri >= len(rows):
            continue
        for col_idx in range(1, min(5, len(rows[ri]))):
            cell = rows[ri][col_idx]
            if not cell or "RS" not in cell:
                continue
            parts = re.split(r"(?=RS[A-Za-z ]+?:)", cell)
            for part in parts:
                part = part.strip()
                if not part.startswith("RS"):
                    continue
                m = re.match(r"(RS[^:]{3,80}):\s*(.+)", part, re.S)
                if not m:
                    continue
                name = m.group(1).strip()
                note = m.group(2).strip().replace("\n", " ")
                product, base = detect_product(note)
                pipeline, _ = detect_pipeline(note)
                val = estimate_value(note, pipeline, product, base)
                out_rows.append(make_row(meta3, {
                    "date": "2026-06-24",
                    "name": name,
                    "city": "",
                    "visit": detect_visit_type(note, pipeline),
                    "product": product,
                    "pipeline": pipeline,
                    "contact": "",
                    "note": note[:500],
                    "estimatedValue": val,
                }, rid3))
    return finalize_rows(out_rows)


def parse_bagus(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows = []
    for sn in wb.sheetnames:
        all_rows.extend(parse_bagus_sheet(wb[sn], sn))
    return all_rows


def build_template_rows() -> list[list]:
    example = [
        "rpt_example_001", "hatim", "Ahmad Hatim Ashshidiq", "2026-06-26", "Minggu 4", 5, 3, "Jateng Utara & Pantura",
        3, 18500, "RS Banyumanik 2",
        "Follow up penawaran C-Arm & CT trade-in",
        "Vendor harus terdaftar formularium Hermina pusat",
        "Visit RS Charlie Demak & Sultan Agung",
        2,
        "RS Banyumanik 2", "Semarang", "followup", "C-Arm", "hot", "Pak Fadil", "Butuh C-Arm dan trade-in CT Scan 64 slice", 4500,
    ]
    plan_row = [
        "rpt_example_002", "icha", "Ika Apriani", "2026-06-29", "Minggu 1", 4, 2, "Jabodetabek & Banten",
        2, 12000, "Royal Taruma Hospital",
        "", "", "Kunjungan W1 Juli Jabodetabek",
        2,
        "Royal Taruma Hospital", "Jakarta Barat", "first", "CT Scan", "warm", "", "Rencana kunjungan", 6000,
    ]
    return [example, plan_row]


def write_xlsx_template(path: Path):
    wb = openpyxl.Workbook()
    guide = wb.active
    guide.title = "PANDUAN"
    guide.append(["PANDUAN TEMPLATE LAPORAN LAPANGAN HNTI → IMS"])
    guide.append([])
    guide.append(["Kolom wajib per baris kunjungan RS:"])
    guide.append(["Report ID", "Sama untuk semua baris dalam 1 laporan mingguan"])
    guide.append(["Sales ID", "hatim | astrika | dwi | icha | tri | bagus | office"])
    guide.append(["Wilayah (ref. peta HNTI)", "Hatim=Jateng Utara/Pantura · Astrika=Jateng Selatan+DIY (Agustus) · Tri=Jatim Sel/Tim · Bagus=Jatim Ut/Bar · Dwi/Ika=Jabodetabek"])
    guide.append(["Date", "Format YYYY-MM-DD"])
    guide.append(["Visit Type", "first | followup | demo | nego | closed"])
    guide.append(["Pipeline Temp", "cold | warm | hot | proposal | win"])
    guide.append(["Estimasi Nilai Proyek (Rp juta)", "Nilai per RS — total otomatis ke Pipeline Value"])
    guide.append([])
    guide.append(["1 baris = 1 kunjungan RS. Report ID sama = 1 laporan mingguan."])

    tpl = wb.create_sheet("TEMPLATE_LAPORAN")
    tpl.append(HEADER)
    for row in build_template_rows():
        tpl.append(row)

    ref = wb.create_sheet("REFERENSI_SALES")
    ref.append(["Sales ID", "Nama", "Wilayah", "Basis", "Catatan"])
    for key in ("hatim", "astrika", "dwi", "icha", "tri", "bagus"):
        sid, name, area = SALES[key]
        basis = {"hatim": "Semarang", "astrika": "Yogyakarta", "dwi": "Jakarta", "icha": "Jakarta", "tri": "Sidoarjo", "bagus": "Surabaya"}.get(key, "")
        note = "Aktif Agustus 2026" if key == "astrika" else ("Under Dwi" if key == "icha" else "")
        ref.append([sid, name, area, basis, note])

    wb.save(path)


def main():
    sources = [
        ("hatim_report", parse_hatim_table(DESKTOP / "PLANING VISIT SALES HATIM 22062026-26062026.xlsx", False)),
        ("hatim_plan", parse_hatim_table(DESKTOP / "PLANING VISIT SALES HATIM 29062026-03072026.xlsx", True)),
        ("tri_report", parse_tri_report(DESKTOP / "DOC-20260628-WA0022..xlsx")),
        ("tri_plan", parse_tri_plan(DESKTOP / "DOC-20260628-WA0021..xlsx")),
        ("ika_plan", parse_ika_plan(DESKTOP / "WEEKLY  SALES PLANNING JULI.xlsx")),
        ("dwi_ika_report", parse_dwi_ika_report(DESKTOP / "REPORT SALES JUNI.xlsx")),
        ("dwi_plan", parse_dwi_plan(DESKTOP / "Dwi Plan 29 Juni-3 Juli.xlsx")),
        ("bagus", parse_bagus(DESKTOP / "PLAN VISIT HNTI_BAGUS.xlsx")),
    ]

    all_rows = []
    for label, rows in sources:
        out = OUT_DIR / f"HNTI_Import_{label}.csv"
        write_csv(out, rows)
        print(f"{label}: {len(rows)} visit rows → {out.name}")
        all_rows.extend(rows)

    write_csv(OUT_DIR / "HNTI_Import_Sales_Reports_All.csv", all_rows)
    write_csv(OUT_DIR / "HNTI_Template_Weekly_Field_Report_IMS.csv", build_template_rows())
    write_xlsx_template(OUT_DIR / "HNTI_Template_Weekly_Field_Report_IMS.xlsx")
    print(f"\nCombined: {len(all_rows)} rows → HNTI_Import_Sales_Reports_All.csv")
    print("Template updated → HNTI_Template_Weekly_Field_Report_IMS.csv / .xlsx")


if __name__ == "__main__":
    main()
